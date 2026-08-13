"""Parsers for DOCX, XLSX, CSV, and Markdown materials.

Each parser returns the same canonical contract as the PDF/image path
(``ParsedOutput`` -> one ``PageResult`` -> ``BlockResult`` -> ``CellResult``)
and attaches a format-native :class:`Locator` to every block and table cell.
Structured formats never invent page numbers: their single page has
``number=None`` and no coordinates.
"""

import csv
import io
import re
from datetime import date, datetime, time
from pathlib import Path
from tempfile import NamedTemporaryFile

import docx
import openpyxl
from docx.oxml.ns import qn
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph as DocxParagraph
from openpyxl.utils import get_column_letter

from app.parsing import BlockResult, CellResult, Locator, PageResult, ParsedOutput, copy_stream

STRUCTURED_EXTENSIONS = {".docx", ".xlsx", ".csv", ".md", ".markdown"}

HEADING_RE = re.compile(r"^(#{1,6})\s+(.*?)\s*#*\s*$")
SETEXT_RE = re.compile(r"^\s*(=+|-+)\s*$")
TABLE_SEPARATOR_RE = re.compile(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)*\|?\s*$")


def decode_text(content: bytes) -> tuple[str, str]:
    """Decode CSV/Markdown bytes, returning (encoding, text).

    UTF-8 (with optional BOM) is tried first; GB18030 covers the common
    Chinese spreadsheet exports that Excel saves without a BOM.
    """
    if content.startswith(b"\xef\xbb\xbf"):
        return "utf-8", content[len(b"\xef\xbb\xbf") :].decode("utf-8")
    try:
        return "utf-8", content.decode("utf-8")
    except UnicodeDecodeError:
        return "gb18030", content.decode("gb18030")


def spreadsheet_text(value) -> str:
    """Render an XLSX cell value the way Excel displays it, without rounding amounts."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return str(value)


def parse_structured(filename: str, stream) -> ParsedOutput:
    extension = Path(filename).suffix.lower()
    if extension == ".docx":
        with NamedTemporaryFile(suffix=extension) as material:
            copy_stream(stream, material)
            blocks = parse_docx(material.name)
        parser_version = f"python-docx-{docx.__version__}"
    elif extension == ".xlsx":
        with NamedTemporaryFile(suffix=extension) as material:
            copy_stream(stream, material)
            blocks = parse_xlsx(material.name)
        parser_version = f"openpyxl-{openpyxl.__version__}"
    elif extension == ".csv":
        blocks = parse_csv(stream.read())
        parser_version = "icrm-csv-1"
    elif extension in {".md", ".markdown"}:
        blocks = parse_markdown(decode_text(stream.read())[1])
        parser_version = "icrm-markdown-1"
    else:
        raise ValueError("unsupported_parser_format")
    return ParsedOutput(
        parser_version=parser_version,
        model_version="none",
        status="success",
        pages=(PageResult(number=None, width=None, height=None, status="success", blocks=blocks),),
    )


def parse_docx(path: str) -> tuple[BlockResult, ...]:
    document = docx.Document(path)
    blocks: list[BlockResult] = []
    for child_index, child in enumerate(document.element.body.iterchildren(), start=1):
        if child.tag == qn("w:p"):
            paragraph = DocxParagraph(child, document)
            text = _docx_paragraph_text(paragraph)
            if not text:
                continue
            style = paragraph.style.name if paragraph.style is not None else ""
            is_heading = style.startswith("Heading") or style.startswith("标题")
            kind = "heading" if is_heading else "paragraph"
            blocks.append(
                BlockResult(
                    len(blocks),
                    kind,
                    text,
                    None,
                    "docx_text",
                    locator=Locator(kind="docx", paragraph_path=f"body/{child_index}"),
                )
            )
        elif child.tag == qn("w:tbl"):
            table = DocxTable(child, document)
            rows = [[cell.text for cell in row.cells] for row in table.rows]
            width = max((len(row) for row in rows), default=0)
            padded = [row + [""] * (width - len(row)) for row in rows]
            cells = tuple(
                CellResult(row_index + 1, column_index + 1, padded[row_index][column_index])
                for row_index in range(len(padded))
                for column_index in range(width)
            )
            if not cells:
                continue
            blocks.append(
                BlockResult(
                    len(blocks),
                    "table",
                    cells[0].text,
                    None,
                    "docx_text",
                    cells=cells,
                    locator=Locator(kind="docx", paragraph_path=f"body/{child_index}"),
                )
            )
    return tuple(blocks)


def _docx_paragraph_text(paragraph: DocxParagraph) -> str:
    parts: list[str] = []
    for node in paragraph._p.iter():
        if node.tag == qn("w:t"):
            parts.append(node.text or "")
        elif node.tag == qn("w:tab"):
            parts.append("\t")
        elif node.tag == qn("w:br"):
            parts.append("\n")
    return "".join(parts).strip()


def parse_xlsx(path: str) -> tuple[BlockResult, ...]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    blocks: list[BlockResult] = []
    for sheet in workbook.worksheets:
        rows = list(
            sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)
        )
        grid = [[spreadsheet_text(cell.value) for cell in row] for row in rows]
        for merged in sheet.merged_cells.ranges:
            text = spreadsheet_text(sheet.cell(merged.min_row, merged.min_col).value)
            for row_index in range(merged.min_row, merged.max_row + 1):
                for column_index in range(merged.min_col, merged.max_col + 1):
                    if row_index - 1 < len(grid) and column_index - 1 < len(grid[row_index - 1]):
                        grid[row_index - 1][column_index - 1] = text
        if not any(value for row in grid for value in row):
            continue
        height = len(grid)
        width = max((len(row) for row in grid), default=0)
        if width == 0:
            continue
        cells = tuple(
            CellResult(
                row_index,
                column_index,
                grid[row_index - 1][column_index - 1],
                locator=Locator(
                    kind="xlsx",
                    sheet=sheet.title,
                    cell=f"{get_column_letter(column_index)}{row_index}",
                ),
            )
            for row_index in range(1, height + 1)
            for column_index in range(1, width + 1)
        )
        blocks.append(
            BlockResult(
                len(blocks),
                "table",
                cells[0].text,
                None,
                "xlsx_text",
                cells=cells,
                locator=Locator(
                    kind="xlsx",
                    sheet=sheet.title,
                    cell_range=f"{get_column_letter(1)}1:{get_column_letter(width)}{height}",
                ),
            )
        )
    return tuple(blocks)


def parse_csv(content: bytes) -> tuple[BlockResult, ...]:
    encoding, text = decode_text(content)
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",\t;").delimiter
    except csv.Error:
        pass
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    while rows and not any(row for row in rows[-1]):
        rows.pop()
    if not rows:
        return ()
    has_header = False
    try:
        has_header = csv.Sniffer().has_header(text[:4096])
    except csv.Error:
        pass
    header = rows[0] if has_header else []
    width = max((len(row) for row in rows), default=0)
    cells = tuple(
        CellResult(
            row_index,
            column_index,
            rows[row_index - 1][column_index - 1]
            if column_index <= len(rows[row_index - 1])
            else "",
            locator=Locator(
                kind="csv",
                row=row_index,
                column=column_index,
                column_name=header[column_index - 1] if column_index <= len(header) else None,
            ),
        )
        for row_index in range(1, len(rows) + 1)
        for column_index in range(1, width + 1)
    )
    block = BlockResult(
        0,
        "table",
        cells[0].text if cells else "",
        None,
        "csv_text",
        cells=cells,
        locator=Locator(kind="csv", row=1, column=1, encoding=encoding),
    )
    return (block,)


def parse_markdown(text: str) -> tuple[BlockResult, ...]:
    lines = text.splitlines()
    blocks: list[BlockResult] = []
    heading_stack: list[tuple[int, str]] = []
    index = 0
    total = len(lines)

    def heading_path() -> str:
        return " / ".join(name for _, name in heading_stack)

    def emit(kind: str, content: str, line_start: int, line_end: int) -> None:
        blocks.append(
            BlockResult(
                len(blocks),
                kind,
                content,
                None,
                "markdown_text",
                locator=Locator(
                    kind="markdown",
                    heading_path=heading_path(),
                    line_start=line_start,
                    line_end=line_end,
                ),
            )
        )

    def push_heading(stack: list[tuple[int, str]], level: int, title: str) -> list[tuple[int, str]]:
        return [(kept_level, name) for kept_level, name in stack if kept_level < level] + [
            (level, title)
        ]

    while index < total:
        line = lines[index]
        stripped = line.strip()
        if not stripped:
            index += 1
            continue
        heading = HEADING_RE.match(line)
        if heading:
            level, title = len(heading.group(1)), heading.group(2).strip()
            heading_stack = push_heading(heading_stack, level, title)
            emit("heading", title, index + 1, index + 1)
            index += 1
            continue
        if stripped.startswith("```") or stripped.startswith("~~~"):
            fence = stripped[:3]
            content: list[str] = []
            first_line = index + 2
            index += 1
            while index < total and not lines[index].strip().startswith(fence):
                content.append(lines[index])
                index += 1
            if content:
                emit("code", "\n".join(content), first_line, index)
            index += 1
            continue
        if index + 1 < total and "|" in line and TABLE_SEPARATOR_RE.match(lines[index + 1]):
            first_line = index + 1
            rows = [_split_pipe_row(line)]
            index += 2
            last_line = first_line + 1
            while index < total and lines[index].strip() and "|" in lines[index]:
                rows.append(_split_pipe_row(lines[index]))
                last_line = index + 1
                index += 1
            width = max((len(row) for row in rows), default=0)
            padded = [row + [""] * (width - len(row)) for row in rows]
            cells = tuple(
                CellResult(row_index + 1, column_index + 1, padded[row_index][column_index])
                for row_index in range(len(padded))
                for column_index in range(width)
            )
            blocks.append(
                BlockResult(
                    len(blocks),
                    "table",
                    cells[0].text if cells else "",
                    None,
                    "markdown_text",
                    cells=cells,
                    locator=Locator(
                        kind="markdown",
                        heading_path=heading_path(),
                        line_start=first_line,
                        line_end=last_line,
                    ),
                )
            )
            continue
        if index + 1 < total and SETEXT_RE.match(lines[index + 1]):
            level = 1 if lines[index + 1].strip().startswith("=") else 2
            heading_stack = push_heading(heading_stack, level, stripped)
            emit("heading", stripped, index + 1, index + 2)
            index += 2
            continue
        first_line = index + 1
        content = [line]
        index += 1
        while index < total:
            current = lines[index]
            if not current.strip():
                break
            if HEADING_RE.match(current) or current.strip().startswith(("```", "~~~")):
                break
            if "|" in current and index + 1 < total and TABLE_SEPARATOR_RE.match(lines[index + 1]):
                break
            content.append(current)
            index += 1
        emit("paragraph", "\n".join(content), first_line, index)
    return tuple(blocks)


def _split_pipe_row(line: str) -> list[str]:
    return [part.strip().replace("\\|", "|") for part in line.strip().strip("|").split("|")]
